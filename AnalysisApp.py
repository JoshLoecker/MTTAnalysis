import tkinter as tk
import openpyxl
from openpyxl.styles import Font
import sys
from Constants import colors


class Analysis:
    def __init__(self, main, data_file, drug_file, colored_buttons, button_list, font_colors):

        min_width: int = 450
        min_height: int = 75

        plus_x: int = 350
        plus_y: int = 200

        self.main_frame: tk.Tk = main
        self.main_frame.title('Data Analysis')

        self.main_frame.geometry("%sx%s+%s+%s" % (min_width, min_height, plus_x, plus_y))
        self.main_frame.rowconfigure(index=0, weight=1)
        self.main_frame.rowconfigure(index=1, weight=1)
        self.main_frame.columnconfigure(0, weight=1)

        self.data_file: str = data_file
        self.drug_file: str = drug_file
        self.colored_buttons: dict = colored_buttons
        self.button_list: dict = button_list
        self.font_colors: list = font_colors
        self.completed_label: tk.Label = tk.Label(master=self.main_frame, text='Your analysis is complete!\nThis window may be closed.')
        self.message_label: tk.Label = tk.Label(master=self.main_frame)
        self.save_again_button: tk.Button = tk.Button(master=self.main_frame)

        self.data_workbook: openpyxl.workbook.workbook.Workbook = openpyxl.load_workbook(filename=self.data_file, data_only=True)

        self.main_frame.grid()

        self.changing_colors()
        self.analyze_data()
        self.save_workbook()

    def save_workbook(self):

        self.main_frame.title("Save")

        try:
            self.main_frame.geometry("%sx%s+%s+%s" % (450, 75, 350, 200))

            self.data_workbook.save(filename=self.data_file)
            self.message_label.config(text='Save successful!\n'
                                           'You may close this window.')
            self.save_again_button.config(text='Exit', command=sys.exit)

            self.message_label.grid(row=0, column=0)
            self.save_again_button.grid(row=1, column=0, pady=(0, 5), ipadx=4)

        except PermissionError:
            self.message_label.config(text='There was an error saving.\n'
                                           'This is most likely due to not closing the Excel '
                                           'file before attempting to save.\nClose the Data '
                                           'File and try again.')
            self.save_again_button.config(text='Try Again', command=self.save_workbook)

            self.message_label.grid(row=0, column=0)
            self.save_again_button.grid(row=1, column=0, pady=(0, 5))

    def changing_colors(self):

        plum_font: openpyxl.styles.fonts.Font = Font(color=str(colors[0])[1:])
        red_font: openpyxl.styles.fonts.Font = Font(color=str(colors[1])[1:])
        green_font: openpyxl.styles.fonts.Font = Font(color=str(colors[2])[1:])
        magenta_font: openpyxl.styles.fonts.Font = Font(color=str(colors[3])[1:])
        cyan_font: openpyxl.styles.fonts.Font = Font(color=str(colors[4])[1:])
        dark_cyan_font: openpyxl.styles.fonts.Font = Font(color=str(colors[5])[1:])
        black_font: openpyxl.styles.fonts.Font = Font(color='000000')
        underline_font: openpyxl.styles.fonts.Font = Font(underline='single')

        results_by_plate: openpyxl.worksheet.worksheet.Worksheet = self.data_workbook['Results by plate']

        try:
            for button in self.colored_buttons:
                button_color = self.colored_buttons[button]
                for row in results_by_plate.iter_rows(min_row=self.button_list[button][0] + 21,
                                                      max_row=self.button_list[button][0] + 21,
                                                      min_col=self.button_list[button][1] + 2,
                                                      max_col=self.button_list[button][1] + 2):

                    for cell in row:
                        if button_color == self.font_colors[0]:
                            cell.font = plum_font

                        elif button_color == self.font_colors[1]:
                            cell.font = red_font

                        elif button_color == self.font_colors[2]:
                            cell.font = green_font

                        elif button_color == self.font_colors[3]:
                            cell.font = magenta_font

                        elif button_color == self.font_colors[4]:
                            cell.font = cyan_font

                        elif button_color == self.font_colors[5]:
                            cell.font = dark_cyan_font

                        elif button_color == self.font_colors[6]:
                            cell.font = black_font

        except TypeError:
            pass

        # create key on the right side
        results_by_plate["P20"]: str = "Key"

        results_by_plate["P21"]: str = "Sensitive Cell Line"
        results_by_plate["P22"]: str = "Sensitive Negative Control"
        results_by_plate["P23"]: str = "Sensitive Positive Control"

        results_by_plate["P24"]: str = "Resistant Cell Line"
        results_by_plate["P25"]: str = "Resistant Negative Control"
        results_by_plate["P26"]: str = "Resistant Positive Control"

        results_by_plate["P27"]: str = "No treatment done"

        # set font colors
        results_by_plate["P20"].font = underline_font

        results_by_plate["P21"].font = plum_font
        results_by_plate["P22"].font = red_font
        results_by_plate["P23"].font = green_font

        results_by_plate["P24"].font = magenta_font
        results_by_plate["P25"].font = cyan_font
        results_by_plate["P26"].font = dark_cyan_font

        results_by_plate["P27"].font = black_font


    def analyze_data(self):

        # delete old worksheet names that are named "Data Analysis"
        sheet_names: list[str] = self.data_workbook.get_sheet_names()
        for value in sheet_names:
            if "Data Analysis" in value:
                self.data_workbook.remove_sheet(self.data_workbook[value])
        self.data_workbook.create_sheet("Data Analysis")
        data_analysis = self.data_workbook['Data Analysis']
        results_by_plate = self.data_workbook['Results by plate']
        absorbance_values = []

        data_analysis['A1'] = 'Data analyzed by a program written in Python by Josh Loecker ' \
                              'using the same protocols originally done by hand'

        for row in results_by_plate.iter_rows(min_row=21, max_row=28, min_col=2, max_col=13):
            for cell in row:
                absorbance_values.append(cell.value)

        row = 5
        column = "A"
        for value in absorbance_values:
            write_to_cell = column + str(row)
            data_analysis[write_to_cell] = value
            row += 1


if __name__ == '__main__':
    root = tk.Tk()
    Analysis(main=root, data_file=None, drug_file=None, colored_buttons=None, button_list=None, font_colors=None)
    root.mainloop()
